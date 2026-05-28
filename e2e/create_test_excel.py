import openpyxl
import os

test_data_dir = "e:/Qingtou_V6/e2e/test-data"
os.makedirs(test_data_dir, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "导入数据"

ws.append(["箱号", "箱状态", "货主", "箱型", "封号"])
ws.append(["TEST0000001", "heavy", "测试客户A", "20GP", "SEAL001"])
ws.append(["TEST0000002", "empty", "测试客户B", "40GP", "SEAL002"])
ws.append(["TEST0000003", "heavy", "测试客户C", "40HQ", "SEAL003"])
ws.append(["TEST0000004", "", "测试客户D", "45HQ", "SEAL004"])
ws.append(["TEST0000005", "", "测试客户E", "20GP", "SEAL005"])

wb_invalid = openpyxl.Workbook()
ws_invalid = wb_invalid.active
ws_invalid.title = "导入数据"

ws_invalid.append(["箱号", "箱状态", "货主", "箱型", "封号"])
ws_invalid.append(["INVALID", "heavy", "测试客户", "20GP", "SEAL001"])
ws_invalid.append(["TEST0000006", "invalid_status", "测试客户", "40GP", "SEAL002"])

wb_partial = openpyxl.Workbook()
ws_partial = wb_partial.active
ws_partial.title = "导入数据"

ws_partial.append(["箱号", "箱状态", "货主", "箱型", "封号"])
ws_partial.append(["TEST0000007", "heavy", "测试客户A", "20GP", "SEAL007"])
ws_partial.append(["INVALID2", "empty", "测试客户B", "40GP", "SEAL008"])
ws_partial.append(["TEST0000008", "heavy", "测试客户C", "40HQ", "SEAL009"])

valid_file = os.path.join(test_data_dir, "valid_import.xlsx")
invalid_file = os.path.join(test_data_dir, "invalid_import.xlsx")
partial_file = os.path.join(test_data_dir, "partial_import.xlsx")

wb.save(valid_file)
wb_invalid.save(invalid_file)
wb_partial.save(partial_file)

print(f"✅ 创建测试文件成功:")
print(f"  - {valid_file} (5条有效数据)")
print(f"  - {invalid_file} (2条无效数据)")
print(f"  - {partial_file} (2条有效，1条无效)")
